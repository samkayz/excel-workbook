from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 4)
        corrected_value = cell.value * 100
        corrected_value_cell = sheet.cell(row, 6)
        corrected_value_cell.value = corrected_value

        values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=6, max_col=6)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'g2')
        
        wb.save(filename)
